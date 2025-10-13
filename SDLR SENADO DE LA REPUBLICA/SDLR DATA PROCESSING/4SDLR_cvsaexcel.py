import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

IN_CSV  = "SDLR RELACION.csv"
OUT_XLS = " SDLR Archivo Relacional.xlsx"

df = pd.read_csv(IN_CSV, dtype=str).fillna("")

wb = Workbook()
ws = wb.active
ws.title = "Reporte"

headers = ["Nombre y Fecha del Documento", "Link pdf", "Enlace youtube"]
ws.append(headers)

# Estilos base
header_fill = PatternFill("solid", fgColor="1F497D")
header_font = Font(bold=True, color="FFFFFF")
align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="DDDDDD")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

# Escribir filas
for _, r in df.iterrows():
    nombre = r["nombre_documento"]
    pdf = r["link_pdf"]
    yt  = r["enlace_youtube"]

    ws.append([nombre, pdf, yt])
    i = ws.max_row
    # Hipervínculos (el texto es la URL completa)
    if pdf:
        ws.cell(i, 2).hyperlink = pdf
        ws.cell(i, 2).style = "Hyperlink"
    if yt:
        ws.cell(i, 3).hyperlink = yt
        ws.cell(i, 3).style = "Hyperlink"

# Aplicar estilos por celda (bordes, alineación, zebra)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
    r = row[0].row
    for c in row:
        c.alignment = align
        c.border = border
        if r == 1:
            c.fill = header_fill
            c.font = header_font
        else:
            # zebra: filas pares con gris claro
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F7F7F7")

# Filtro y congelar encabezado
ws.auto_filter.ref = f"A1:C{ws.max_row}"
ws.freeze_panes = "A2"

# Auto-ancho de columnas (cap a 120 “caracteres” aprox)
for col in range(1, 4):
    letter = get_column_letter(col)
    max_len = len(headers[col-1])
    for cell in ws[letter]:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[letter].width = min(max_len + 2, 120)

# Altura de filas para ver URLs largas cómodas
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 28

# Crear “tabla” con estilo (mantiene filtros y zebra propios)
tbl = Table(displayName="TablaReporte", ref=f"A1:C{ws.max_row}")
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False
)
ws.add_table(tbl)

wb.save(OUT_XLS)
print(f"✅ Listo: {OUT_XLS}")

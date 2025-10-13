# 2institucionesexcel.py
from pathlib import Path
from openpyxl import load_workbook, Workbook
from copy import copy
import unicodedata, re

ROOT = Path(__file__).parent  # ...\LLM DATA

def unaccent(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))

def norm(s: str) -> str:
    s = unaccent(s).lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s

def find_relacional(base_dir: Path, must=("archivo","relacional")) -> Path:
    base = ROOT / base_dir
    if not base.exists():
        raise FileNotFoundError(f"No existe carpeta: {base}")
    print(f"📂 En {base} hay:")
    for p in base.iterdir():
        if p.is_file():
            print(" -", repr(p.name))
    for p in base.glob("*.xls*"):
        n = norm(p.name)
        if all(w in n for w in must):
            return p
    raise FileNotFoundError(f"No encontré Excel 'Archivo Relacional' en: {base}")

# Localiza los dos Excels sin depender de tildes/espacios exactos
CDD_PATH  = find_relacional(Path("CDD CAMARA DE DIPUTADOS") / "CDD DATA PROCESSING")
SDLR_PATH = find_relacional(Path("SDLR SENADO DE LA REPUBLICA") / "SDLR DATA PROCESSING")

print("✅ CDD ->", CDD_PATH.resolve())
print("✅ SDLR->", SDLR_PATH.resolve())

def nombre_hoja_desde_ruta(p: Path) -> str:
    return Path(p).parent.parent.name[:31]  # nombre institución (abuelo)

def copiar_hoja(ws_src, ws_dst):
    for k, col_dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[k].width = col_dim.width
    for idx, row_dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[idx].height = row_dim.height
    for row in ws_src.iter_rows():
        for c in row:
            d = ws_dst.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                if c.font:          d.font = copy(c.font)
                if c.fill:          d.fill = copy(c.fill)
                if c.border:        d.border = copy(c.border)
                if c.number_format: d.number_format = c.number_format
                if c.protection:    d.protection = copy(c.protection)
                if c.alignment:     d.alignment = copy(c.alignment)
            if c.hyperlink:
                target = getattr(c.hyperlink, "target", None) or str(c.hyperlink)
                d.hyperlink = target
                d.style = "Hyperlink"
    for r in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(r))
    if ws_src.auto_filter and ws_src.auto_filter.ref:
        ws_dst.auto_filter.ref = ws_src.auto_filter.ref
    if ws_src.freeze_panes:
        ws_dst.freeze_panes = ws_src.freeze_panes

# Construye el archivo combinado
wb_out = Workbook(); wb_out.remove(wb_out.active)
for path in (CDD_PATH, SDLR_PATH):
    wb = load_workbook(path, data_only=False)
    ws_src = wb.active  # usa wb["Reporte"] si prefieres
    ws_dst = wb_out.create_sheet(title=nombre_hoja_desde_ruta(path))
    copiar_hoja(ws_src, ws_dst)

OUT = "ARCHIVOS RELACIONALES.xlsx"
wb_out.save(OUT)
print("OK ->", OUT)
